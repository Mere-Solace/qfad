"""Build/update the master Excel workbook from FRED CSV files.

Usage:
    python scripts/build_master_sheet.py

Strategy:
    - If the workbook doesn't exist, create it from scratch.
    - If it does exist, delete data sheets (All Data, Metadata, per-category)
      but preserve any user-created sheets (Charts, Analysis, etc.).
    - Recreate data sheets with fresh data from CSVs.

Sheets created:
    - "All Data"          — every series outer-joined on date, forward-filled
    - Per-category sheets — Treasury Rates, Spreads, Inflation, etc.
    - "Metadata"          — series ID, name, unit, frequency, category, last date, rows
"""

import json
import logging
import sys
from datetime import date
from pathlib import Path

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows

PROJECT_ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(PROJECT_ROOT))

logging.basicConfig(level=logging.INFO, format="%(asctime)s  %(message)s", datefmt="%H:%M:%S")
log = logging.getLogger(__name__)

SERIES_CONFIG = PROJECT_ROOT / "config" / "series.json"
DATA_DIR = PROJECT_ROOT / "data" / "raw" / "fred"
WORKBOOK_PATH = PROJECT_ROOT / "data" / "master_workbook.xlsx"

# Category key -> display name for sheet tabs
CATEGORY_DISPLAY = {
    "treasury_rates": "Treasury Rates",
    "spreads": "Spreads",
    "inflation": "Inflation",
    "employment": "Employment",
    "leading_indicators": "Leading Indicators",
    "monetary_policy": "Monetary Policy",
    "financial_conditions": "Financial Conditions",
    "output": "Output",
    "recession_indicators": "Recession Indicators",
}

# Sheet names managed by this script (will be deleted and recreated)
MANAGED_SHEETS = {"All Data", "All Data Monthly", "Metadata"} | set(CATEGORY_DISPLAY.values())


def load_config() -> dict:
    """Load series config, returning the fred dict keyed by category."""
    with open(SERIES_CONFIG) as f:
        return json.load(f).get("fred", {})


def load_all_csvs(config: dict) -> dict[str, pd.DataFrame]:
    """Load all CSV files into a dict of series_id -> DataFrame."""
    series_data = {}
    for _cat, items in config.items():
        for item in items:
            sid = item["id"]
            path = DATA_DIR / f"{sid}.csv"
            if path.exists():
                df = pd.read_csv(path, parse_dates=["date"])
                df["date"] = pd.to_datetime(df["date"]).dt.date
                df = df.sort_values("date").reset_index(drop=True)
                series_data[sid] = df
    return series_data


def build_all_data_sheet(series_data: dict[str, pd.DataFrame], config: dict) -> pd.DataFrame:
    """Build the 'All Data' wide DataFrame: date | series1 | series2 | ..."""
    if not series_data:
        return pd.DataFrame()

    # Build a full date index spanning all series
    all_dates: set[date] = set()
    for df in series_data.values():
        all_dates.update(df["date"].tolist())

    date_index = sorted(all_dates)
    result = pd.DataFrame({"date": date_index})

    # Build a lookup from series_id to display name
    name_map = {}
    for _cat, items in config.items():
        for item in items:
            name_map[item["id"]] = item.get("name", item["id"])

    for sid, df in series_data.items():
        renamed = df.rename(columns={"value": sid})
        result = result.merge(renamed[["date", sid]], on="date", how="left")

    # Forward-fill to align different frequencies, then sort newest-first
    result = result.sort_values("date").reset_index(drop=True)
    result.iloc[:, 1:] = result.iloc[:, 1:].ffill()
    result = result.sort_values("date", ascending=False).reset_index(drop=True)

    return result


def build_category_sheet(
    series_data: dict[str, pd.DataFrame], items: list[dict]
) -> pd.DataFrame:
    """Build a per-category wide DataFrame."""
    sids = [item["id"] for item in items if item["id"] in series_data]
    if not sids:
        return pd.DataFrame()

    all_dates: set[date] = set()
    for sid in sids:
        all_dates.update(series_data[sid]["date"].tolist())

    date_index = sorted(all_dates)
    result = pd.DataFrame({"date": date_index})

    for sid in sids:
        renamed = series_data[sid].rename(columns={"value": sid})
        result = result.merge(renamed[["date", sid]], on="date", how="left")

    # Forward-fill to align different frequencies, then sort newest-first
    result = result.sort_values("date").reset_index(drop=True)
    result.iloc[:, 1:] = result.iloc[:, 1:].ffill()
    result = result.sort_values("date", ascending=False).reset_index(drop=True)

    return result


def build_metadata(series_data: dict[str, pd.DataFrame], config: dict) -> pd.DataFrame:
    """Build the metadata DataFrame."""
    rows = []
    for cat_key, items in config.items():
        display_cat = CATEGORY_DISPLAY.get(cat_key, cat_key)
        for item in items:
            sid = item["id"]
            df = series_data.get(sid)
            rows.append({
                "series_id": sid,
                "name": item.get("name", sid),
                "unit": item.get("unit", ""),
                "frequency": item.get("frequency", ""),
                "category": display_cat,
                "last_date": str(max(df["date"])) if df is not None and not df.empty else "",
                "rows": len(df) if df is not None else 0,
            })
    return pd.DataFrame(rows)


def _build_description_map(config: dict) -> dict[str, str]:
    """Build a map of series_id -> 'Name (unit)' for the description row."""
    desc = {}
    for _cat, items in config.items():
        for item in items:
            name = item.get("name", item["id"])
            unit = item.get("unit", "")
            desc[item["id"]] = f"{name} ({unit})" if unit else name
    return desc


def write_df_to_sheet(
    wb: Workbook,
    sheet_name: str,
    df: pd.DataFrame,
    desc_map: dict[str, str] | None = None,
):
    """Write a DataFrame to a new sheet in the workbook.

    If desc_map is provided, inserts a description row (row 1) above the
    header row with the series name and units for each column.
    """
    ws = wb.create_sheet(title=sheet_name)
    start_row = 1

    if desc_map:
        # Row 1: descriptions
        ws.cell(row=1, column=1, value="")
        for c_idx, col_name in enumerate(df.columns[1:], 2):
            ws.cell(row=1, column=c_idx, value=desc_map.get(col_name, ""))
        start_row = 2

    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), start_row):
        for c_idx, value in enumerate(row, 1):
            ws.cell(row=r_idx, column=c_idx, value=value)

    # Set date column width to 14 (~164 px)
    ws.column_dimensions["A"].width = 14


def build_monthly_df(all_data_df: pd.DataFrame) -> pd.DataFrame:
    """Downsample the All Data DataFrame to the first available date per month."""
    if all_data_df.empty:
        return pd.DataFrame()

    df = all_data_df.copy()
    # Sort oldest-first so groupby keeps the earliest date in each month
    df = df.sort_values("date").reset_index(drop=True)
    df["_ym"] = df["date"].apply(lambda d: (d.year, d.month))
    monthly = df.groupby("_ym", sort=True).first().reset_index(drop=True)
    # Sort newest-first to match All Data ordering
    monthly = monthly.sort_values("date", ascending=False).reset_index(drop=True)
    return monthly


def build_master_sheet():
    """Build or update the master Excel workbook."""
    config = load_config()
    series_data = load_all_csvs(config)
    log.info("Loaded %d series from CSV", len(series_data))

    # Open existing workbook or create new
    if WORKBOOK_PATH.exists():
        wb = load_workbook(WORKBOOK_PATH)
        # Remove managed sheets (preserve user sheets)
        for name in list(wb.sheetnames):
            if name in MANAGED_SHEETS:
                del wb[name]
        log.info("Opened existing workbook, preserved user sheets: %s",
                 [s for s in wb.sheetnames if s not in MANAGED_SHEETS])
    else:
        wb = Workbook()
        # Remove the default empty sheet
        wb.remove(wb.active)
        WORKBOOK_PATH.parent.mkdir(parents=True, exist_ok=True)

    desc_map = _build_description_map(config)

    # All Data sheet
    all_data_df = build_all_data_sheet(series_data, config)
    if not all_data_df.empty:
        write_df_to_sheet(wb, "All Data", all_data_df, desc_map)
        log.info("All Data: %d rows x %d columns", len(all_data_df), len(all_data_df.columns))

    # All Data Monthly sheet (first date per month)
    monthly_df = build_monthly_df(all_data_df)
    if not monthly_df.empty:
        write_df_to_sheet(wb, "All Data Monthly", monthly_df, desc_map)
        log.info("All Data Monthly: %d rows x %d columns", len(monthly_df), len(monthly_df.columns))

    # Per-category sheets
    for cat_key, items in config.items():
        display_name = CATEGORY_DISPLAY.get(cat_key, cat_key)
        cat_df = build_category_sheet(series_data, items)
        if not cat_df.empty:
            write_df_to_sheet(wb, display_name, cat_df, desc_map)
            log.info("%s: %d rows x %d columns", display_name, len(cat_df), len(cat_df.columns))

    # Metadata sheet
    meta_df = build_metadata(series_data, config)
    write_df_to_sheet(wb, "Metadata", meta_df)

    wb.save(WORKBOOK_PATH)
    log.info("Saved workbook to %s", WORKBOOK_PATH)


if __name__ == "__main__":
    build_master_sheet()
